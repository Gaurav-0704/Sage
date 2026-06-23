"""
Shared tabular I/O for the round-trip Excel/CSV sync (STEP 2).

One reader (`read_tabular`) turns an uploaded .xlsx OR .csv into a list of
``{header: value}`` string dicts, so the existing per-domain `_row_to_fields`
validators work unchanged for both formats. One writer (`build_xlsx`) renders
rows back into a .xlsx using the *exact same column layout* as the import
template, so export → edit → import round-trips cleanly.

openpyxl is the only Excel dependency; it stays isolated in this module.
"""

import csv
import io

from openpyxl import Workbook, load_workbook

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _stringify(v) -> str:
    """Render any cell value as the trimmed string our validators expect."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "yes" if v else "no"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    iso = getattr(v, "isoformat", None)  # date / datetime -> ISO
    if callable(iso) and not isinstance(v, str):
        return iso()
    return str(v).strip()


def read_tabular(filename: str, raw: bytes) -> list[dict]:
    """Parse an uploaded .csv or .xlsx into a list of header->value dicts.

    Raises ValueError on an unsupported format or a missing header row.
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _read_xlsx(raw)
    if name.endswith(".csv"):
        return _read_csv(raw)
    raise ValueError("Unsupported file type — upload a .xlsx or .csv file")


def _read_csv(raw: bytes) -> list[dict]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("File is empty or missing a header row")
    return [dict(r) for r in reader]


def _read_xlsx(raw: bytes) -> list[dict]:
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — surface a clean message to the API
        raise ValueError(f"Could not read .xlsx file: {e}")
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    try:
        header = next(row_iter)
    except StopIteration:
        wb.close()
        raise ValueError("File is empty or missing a header row")

    headers = [_stringify(h) for h in header]
    if not any(headers):
        wb.close()
        raise ValueError("File is empty or missing a header row")

    out: list[dict] = []
    for r in row_iter:
        if r is None or all(c is None or _stringify(c) == "" for c in r):
            continue  # skip fully-blank rows
        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = _stringify(r[i]) if i < len(r) else ""
        out.append(d)
    wb.close()
    return out


def build_xlsx(columns: list[str], rows: list[dict], sheet_title: str = "Sheet1") -> bytes:
    """Render rows into a .xlsx workbook with `columns` as the header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Sheet1")[:31]
    ws.append(list(columns))
    for row in rows:
        ws.append([_stringify(row.get(c, "")) for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
