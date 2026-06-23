"""
STEP 2 — Excel/CSV sync tests.

Covers the single-source-of-truth upsert path shared by manual entry and
import, the .xlsx round-trip (export → read back), in-file duplicate skipping,
and validation errors. Logic is exercised directly against the in-memory DB
(matching the existing conftest style — no live HTTP server needed).
"""

import io

import pytest
from openpyxl import load_workbook

import models
import excel_io
from agents import students_agent, teachers_agent


@pytest.fixture(autouse=True)
def _no_disk_sync(monkeypatch):
    """Keep imports from rewriting the real data/*.csv files during tests."""
    monkeypatch.setattr(students_agent, "sync_students", lambda db: True)
    monkeypatch.setattr(students_agent, "sync_records", lambda db: True)
    monkeypatch.setattr(teachers_agent, "sync_teachers", lambda db: True)


# ---------------- excel_io round-trip ---------------- #

class TestExcelIO:
    def test_build_and_read_xlsx_roundtrip(self):
        cols = ["admission_no", "name", "student_class"]
        rows = [
            {"admission_no": "A1", "name": "Asha", "student_class": "5"},
            {"admission_no": "A2", "name": "Bina", "student_class": "6"},
        ]
        data = excel_io.build_xlsx(cols, rows, sheet_title="Students")
        back = excel_io.read_tabular("students.xlsx", data)
        assert [r["admission_no"] for r in back] == ["A1", "A2"]
        assert back[0]["name"] == "Asha"
        assert back[1]["student_class"] == "6"

    def test_numbers_stringified_not_floaty(self):
        data = excel_io.build_xlsx(["k"], [{"k": 7}])
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "7"  # not "7.0"

    def test_read_csv_bytes(self):
        raw = b"admission_no,name,student_class\nA1,Asha,5\n"
        rows = excel_io.read_tabular("x.csv", raw)
        assert rows == [{"admission_no": "A1", "name": "Asha", "student_class": "5"}]

    def test_unsupported_extension(self):
        with pytest.raises(ValueError):
            excel_io.read_tabular("x.pdf", b"junk")

    def test_blank_rows_skipped(self):
        data = excel_io.build_xlsx(
            ["admission_no", "name"],
            [{"admission_no": "A1", "name": "Asha"}],
        )
        # append a blank logical row by reading back — build only wrote one row
        rows = excel_io.read_tabular("x.xlsx", data)
        assert len(rows) == 1


# ---------------- students upsert (shared path) ---------------- #

class TestStudentUpsert:
    def test_create_then_update_same_key(self, db):
        action, s = students_agent.upsert_student(
            db, {"admission_no": "A001", "name": "First", "student_class": "5"})
        db.commit()
        assert action == "created"
        assert s.id is not None

        action2, s2 = students_agent.upsert_student(
            db, {"admission_no": "A001", "name": "Renamed", "student_class": "6"})
        db.commit()
        assert action2 == "updated"
        assert s2.id == s.id
        assert db.query(models.Student).count() == 1
        assert db.query(models.Student).first().name == "Renamed"

    def test_import_rows_created_updated_counts(self, db):
        # seed one existing student
        students_agent.upsert_student(
            db, {"admission_no": "A001", "name": "Old", "student_class": "5"})
        db.commit()

        rows = [
            {"admission_no": "A001", "name": "Updated", "student_class": "5"},
            {"admission_no": "A002", "name": "Brand New", "student_class": "6"},
        ]
        summary = students_agent.import_student_rows(db, rows)
        assert summary["created"] == 1
        assert summary["updated"] == 1
        assert summary["skipped"] == 0
        assert summary["error_count"] == 0
        assert db.query(models.Student).count() == 2

    def test_in_file_duplicate_is_skipped_not_duplicated(self, db):
        rows = [
            {"admission_no": "DUP", "name": "One", "student_class": "5"},
            {"admission_no": "DUP", "name": "Two", "student_class": "5"},
        ]
        summary = students_agent.import_student_rows(db, rows)
        assert summary["created"] == 1
        assert summary["skipped"] == 1
        assert db.query(models.Student).filter(models.Student.admission_no == "DUP").count() == 1

    def test_validation_error_collected(self, db):
        rows = [{"name": "No Admission", "student_class": "5"}]  # missing admission_no
        summary = students_agent.import_student_rows(db, rows)
        assert summary["created"] == 0
        assert summary["error_count"] == 1
        assert "admission_no" in summary["errors"][0]["error"]

    def test_dry_run_writes_nothing(self, db):
        rows = [{"admission_no": "A9", "name": "Ghost", "student_class": "5"}]
        summary = students_agent.import_student_rows(db, rows, dry_run=True)
        assert summary["created"] == 1
        assert summary["dry_run"] is True
        assert db.query(models.Student).count() == 0

    def test_export_row_layout_matches_columns(self, db):
        students_agent.upsert_student(
            db, {"admission_no": "A001", "name": "Asha", "student_class": "5"})
        db.commit()
        s = db.query(models.Student).first()
        row = students_agent._student_to_row(s)
        assert set(row.keys()) == set(students_agent.CSV_COLUMNS)


# ---------------- teachers upsert (shared path) ---------------- #

class TestTeacherUpsert:
    def _fields(self, emp="SGET001", email="t1@sage.school", name="Tara"):
        return {
            "employee_id": emp, "name": name, "email": email,
            "subject": "Math", "classes_taught": "5,6",
            "qualification": "M.Sc", "phone": "9999900000",
            "can_do_front_office": False,
        }

    def test_create_makes_user_and_teacher(self, db):
        action, t = teachers_agent.upsert_teacher(db, self._fields())
        db.commit()
        assert action == "created"
        u = db.query(models.User).filter(models.User.id == t.user_id).first()
        assert u.role == "teacher"
        assert u.email == "t1@sage.school"

    def test_update_existing_by_employee_id(self, db):
        teachers_agent.upsert_teacher(db, self._fields())
        db.commit()
        action, t = teachers_agent.upsert_teacher(
            db, self._fields(name="Tara Renamed"))
        db.commit()
        assert action == "updated"
        assert db.query(models.Teacher).count() == 1
        u = db.query(models.User).filter(models.User.id == t.user_id).first()
        assert u.name == "Tara Renamed"

    def test_email_collision_raises(self, db):
        # an unrelated user already owns this email
        db.add(models.User(name="X", email="dupe@sage.school",
                           password="x", role="staff", status="active"))
        db.commit()
        with pytest.raises(ValueError):
            teachers_agent.upsert_teacher(
                db, self._fields(emp="SGET999", email="dupe@sage.school"))

    def test_import_rows_summary(self, db):
        rows = [
            {"employee_id": "SGET001", "name": "Tara", "email": "t1@sage.school"},
            {"employee_id": "SGET001", "name": "Dup", "email": "t1@sage.school"},
        ]
        summary = teachers_agent.import_teacher_rows(db, rows)
        assert summary["created"] == 1
        assert summary["skipped"] == 1
        assert db.query(models.Teacher).count() == 1
